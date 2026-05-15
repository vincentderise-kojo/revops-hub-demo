"""Assign master accounts to territories by nearest marquee distance.

Reads:
  - inputs/marquees/<TerritoryName>.csv       one CSV per territory (filename = label),
                                              OR a single CSV with a Territory column,
                                              OR a single CSV with no Territory column +
                                              inputs/territory_mapping.csv (Account Name → Territory)
  - --master <path>                           classified master xlsx (or raw if --include-all)
  - --enrich-master <path>                    optional fresh SFDC pull (CSV/xlsx) merged onto
                                              classified master by Account ID (18 Char) — used to
                                              bring in Parent Account ID, Sales Last Activity Date,
                                              and a refreshed Account Owner without re-classifying.

Writes:
  - inputs/assigned_<ts>.xlsx with sheets:
      • Assigned Accounts   — per-account list with Territory + flags (uses primary cap)
      • Summary             — per-territory roll-up with flag counts
      • Territory Marquees  — Territory → marquee anchor reference

Logic:
  1. Geocode every ZIP via pgeocode (US Census GeoNames data).
  2. For each territory, distance = min(distance to any marquee in that territory).
  3. Pick top-N classifier-confirmed ICP Fits per territory within max-distance cap.
  4. Resolve overlaps: each account assigned to its single closest territory;
     other claiming territories listed in Secondary Territory.
  5. Compute per-account flags from SFDC fields:
       OpCo?           = Parent Account ID is not null
       Active <14d?    = Sales Last Activity Date within last N days (default 14)
       Has Active Owner? = Account Owner is not a placeholder marketing user
"""

from __future__ import annotations

import argparse
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import pgeocode

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_TARGET = 100
DEFAULT_PRIMARY_CAP_MI = 300  # used for Assigned Accounts sheet + flag counts
COMPARE_CAPS_MI = [200, 300, 500]  # always rendered as columns on Summary
DEFAULT_ACTIVITY_DAYS = 14
EARTH_RADIUS_MI = 3958.7613

# SFDC placeholder owner names that mean "unassigned / open territory."
# Account Owner equal to one of these is treated as no real rep.
MARKETING_USERS = {"Marketing User"}


def _normalize_zip(z) -> str | None:
    """Strip ZIP+4, decimals, whitespace; left-pad to 5 digits. Return None if invalid."""
    if pd.isna(z):
        return None
    s = str(z).strip().split("-")[0].split(".")[0]
    s = s.zfill(5)
    return s if s.isdigit() and len(s) == 5 else None


def load_marquees(marquees_dir: Path, territory_mapping_path: Path | None = None) -> pd.DataFrame:
    """Read all marquee CSVs in directory and resolve their territory labels.

    Territory assignment, in order of precedence:
      1. If territory_mapping_path is provided and exists, join Account Name → Territory.
         Marquees not present in the mapping are skipped with a warning.
      2. Else, if a CSV has a Territory column, use that per-row.
      3. Else, the filename (sans .csv) is used as the Territory label for all rows in that file.
    """
    rows = []
    for csv_path in sorted(marquees_dir.glob("*.csv")):
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        name_col = next((c for c in df.columns if c.strip().lower() == "account name"), None)
        zip_col = next((c for c in df.columns if "zip" in c.lower() or "postal" in c.lower()), None)
        terr_col = next((c for c in df.columns if c.strip().lower() == "territory"), None)
        if not name_col or not zip_col:
            print(f"WARNING: {csv_path.name} missing Account Name or ZIP; skipping", file=sys.stderr)
            continue
        for _, row in df.iterrows():
            rows.append({
                "Marquee Account Name": str(row[name_col]).strip(),
                "Marquee ZIP": _normalize_zip(row[zip_col]),
                "_territory_from_csv": (str(row[terr_col]).strip() if terr_col else None),
                "_filename_territory": csv_path.stem,
            })
    if not rows:
        return pd.DataFrame(columns=["Territory", "Marquee Account Name", "Marquee ZIP"])
    df = pd.DataFrame(rows)

    if territory_mapping_path and territory_mapping_path.exists():
        mapping = pd.read_csv(territory_mapping_path, encoding="utf-8-sig")
        mapping["Account Name"] = mapping["Account Name"].astype(str).str.strip()
        mapping["Territory"] = mapping["Territory"].astype(str).str.strip()
        mapping_dict = dict(zip(mapping["Account Name"], mapping["Territory"]))
        df["Territory"] = df["Marquee Account Name"].map(mapping_dict)
        unmapped = df[df["Territory"].isna()]["Marquee Account Name"].tolist()
        if unmapped:
            print(f"WARNING: {len(unmapped)} marquees not in territory mapping (skipping):", file=sys.stderr)
            for n in unmapped:
                print(f"  - {n}", file=sys.stderr)
        df = df.dropna(subset=["Territory"]).reset_index(drop=True)
    else:
        def pick(r):
            v = r["_territory_from_csv"]
            return v if (v and v.lower() != "nan") else r["_filename_territory"]
        df["Territory"] = df.apply(pick, axis=1)

    return df[["Territory", "Marquee Account Name", "Marquee ZIP"]].reset_index(drop=True)


def merge_enrichment(master_df: pd.DataFrame, enrich_path: Path) -> pd.DataFrame:
    """Merge SFDC enrichment fields onto classified master by Account ID (18 Char).

    Pulls in Parent Account ID, Parent Account, Sales Last Activity Date, and refreshes
    Account Owner. Master rows without a match keep their existing values for any column
    the enrichment doesn't supply.
    """
    if enrich_path.suffix.lower() == ".csv":
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                enrich = pd.read_csv(enrich_path, encoding=enc, dtype=str)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError(f"Could not decode {enrich_path} with utf-8-sig/cp1252/latin-1")
    else:
        enrich = pd.read_excel(enrich_path, dtype=str)

    id_col = "Account ID (18 Char)"
    if id_col not in enrich.columns:
        for alt in ("Account ID", "AccountID"):
            if alt in enrich.columns:
                enrich = enrich.rename(columns={alt: id_col})
                break
        else:
            raise ValueError(f"Enrichment file missing Account ID column. Columns: {list(enrich.columns)}")

    keep = [id_col, "Parent Account ID", "Parent Account", "Sales Last Activity Date", "Account Owner"]
    keep = [c for c in keep if c in enrich.columns]
    enrich_subset = enrich[keep].drop_duplicates(subset=[id_col]).copy()

    # Refresh overlapping columns from enrichment instead of carrying stale master values.
    overlap = [c for c in keep if c != id_col and c in master_df.columns]
    if overlap:
        master_df = master_df.drop(columns=overlap)

    return master_df.merge(enrich_subset, on=id_col, how="left")


def compute_flags(df: pd.DataFrame, activity_days: int = DEFAULT_ACTIVITY_DAYS) -> pd.DataFrame:
    """Add three boolean flag columns to the dataframe (in place safe — returns the df)."""
    df = df.copy()

    if "Parent Account ID" in df.columns:
        pid = df["Parent Account ID"].astype(str).str.strip()
        df["OpCo?"] = pid.notna() & (pid != "") & (pid.str.lower() != "nan")
    else:
        df["OpCo?"] = False

    if "Sales Last Activity Date" in df.columns:
        dt = pd.to_datetime(df["Sales Last Activity Date"], errors="coerce", format="mixed")
        threshold = pd.Timestamp(datetime.now() - timedelta(days=activity_days))
        df["Active <14d?"] = dt.notna() & (dt >= threshold)
    else:
        df["Active <14d?"] = False

    if "Account Owner" in df.columns:
        owner = df["Account Owner"].astype(str).str.strip()
        df["Has Active Owner?"] = ~owner.isin(MARKETING_USERS) & (owner != "") & (owner.str.lower() != "nan")
    else:
        df["Has Active Owner?"] = False

    return df


def geocode_zips(zips: list[str]) -> pd.DataFrame:
    """Look up lat/lon for a list of ZIPs. Returns DataFrame indexed by ZIP."""
    nom = pgeocode.Nominatim("us")
    unique = sorted({z for z in zips if z})
    res = nom.query_postal_code(unique)
    if isinstance(res, pd.Series):
        res = res.to_frame().T
    return res[["postal_code", "latitude", "longitude"]].set_index("postal_code")


def haversine_miles(lat1, lon1, lat2, lon2):
    """Vectorized haversine in miles. Inputs in degrees; supports broadcasting."""
    lat1, lon1, lat2, lon2 = (np.radians(x) for x in (lat1, lon1, lat2, lon2))
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = np.sin(dlat / 2) ** 2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon / 2) ** 2
    return 2 * EARTH_RADIUS_MI * np.arcsin(np.sqrt(a))


def assign(master_df: pd.DataFrame, marquees_df: pd.DataFrame, target: int, max_distance_mi: float):
    """Return (assignments_df, summary_df)."""
    master_df = master_df.copy()
    zip_col = next((c for c in master_df.columns if "zip" in c.lower() or "postal" in c.lower()), None)
    if zip_col is None:
        raise ValueError(f"Master missing ZIP column. Columns: {list(master_df.columns)}")
    master_df["_zip"] = master_df[zip_col].apply(_normalize_zip)

    all_zips = list(set(master_df["_zip"].dropna()) | set(marquees_df["Marquee ZIP"].dropna()))
    geo = geocode_zips(all_zips)

    master_df = master_df.join(geo, on="_zip")
    pre_geo = len(master_df)
    master_df = master_df.dropna(subset=["latitude", "longitude"]).reset_index(drop=True)
    geo_dropped_master = pre_geo - len(master_df)

    marquees_df = marquees_df.join(geo, on="Marquee ZIP", rsuffix="_g")
    pre_geo_m = len(marquees_df)
    marquees_df = marquees_df.dropna(subset=["latitude", "longitude"]).reset_index(drop=True)
    geo_dropped_marquees = pre_geo_m - len(marquees_df)

    if geo_dropped_master or geo_dropped_marquees:
        print(f"  Geocoding dropped {geo_dropped_master} master rows and {geo_dropped_marquees} marquees (missing/unknown ZIP)")

    dist = haversine_miles(
        master_df["latitude"].values[:, None],
        master_df["longitude"].values[:, None],
        marquees_df["latitude"].values[None, :],
        marquees_df["longitude"].values[None, :],
    )

    territories = sorted(marquees_df["Territory"].unique())
    n_master = len(master_df)
    territory_dist = pd.DataFrame(np.full((n_master, len(territories)), np.inf), columns=territories)
    closest_marquee = pd.DataFrame(np.full((n_master, len(territories)), "", dtype=object), columns=territories)

    for terr in territories:
        mask = marquees_df["Territory"].values == terr
        sub_dist = dist[:, mask]
        sub_names = marquees_df.loc[mask, "Marquee Account Name"].values
        min_idx = sub_dist.argmin(axis=1)
        territory_dist[terr] = sub_dist[np.arange(n_master), min_idx]
        closest_marquee[terr] = sub_names[min_idx]

    # Per-territory full eligible lists (sorted nearest first), used for both top-N
    # nomination and shortfall backfill.
    eligible_by_terr: dict[str, list[int]] = {}
    for terr in territories:
        elig = master_df.index[territory_dist[terr] <= max_distance_mi].tolist()
        elig.sort(key=lambda i: territory_dist.at[i, terr])
        eligible_by_terr[terr] = elig

    candidates = []
    for terr, elig in eligible_by_terr.items():
        for i in elig[:target]:
            candidates.append({
                "_idx": i,
                "Territory": terr,
                "Distance to Marquee (mi)": float(territory_dist.at[i, terr]),
                "Closest Marquee": closest_marquee.at[i, terr],
            })
    cand_df = pd.DataFrame(candidates)
    if cand_df.empty:
        result = pd.DataFrame()
    else:
        cand_df = cand_df.sort_values("Distance to Marquee (mi)")
        primary = cand_df.drop_duplicates("_idx", keep="first").copy()

        # Secondary Territory: any other territory whose top-N also nominated this account.
        all_terr = (cand_df.groupby("_idx")["Territory"]
                    .apply(lambda s: sorted(set(s))).rename("_all_terrs"))
        primary = primary.merge(all_terr, on="_idx")
        primary["Secondary Territory"] = primary.apply(
            lambda r: ", ".join(t for t in r["_all_terrs"] if t != r["Territory"]),
            axis=1,
        )
        primary = primary.drop(columns=["_all_terrs"])

        # Backfill: any territory under target walks past position N in its eligible list
        # and claims any account not yet assigned to a primary territory. Naturally preserves
        # no-overlap because we update `claimed` in lock-step. Stops at the territory's
        # in-range pool — won't pull from beyond max_distance_mi.
        claimed = set(primary["_idx"].tolist())
        backfill_rows = []
        for terr in territories:
            count = int((primary["Territory"] == terr).sum())
            needed = target - count
            if needed <= 0:
                continue
            for i in eligible_by_terr[terr][target:]:  # already past the natural top-N
                if i in claimed:
                    continue
                backfill_rows.append({
                    "_idx": i,
                    "Territory": terr,
                    "Distance to Marquee (mi)": float(territory_dist.at[i, terr]),
                    "Closest Marquee": closest_marquee.at[i, terr],
                    "Secondary Territory": "",  # backfilled rows weren't in any other top-N
                })
                claimed.add(i)
                needed -= 1
                if needed == 0:
                    break
        if backfill_rows:
            primary = pd.concat([primary, pd.DataFrame(backfill_rows)], ignore_index=True)

        drop_cols = ["_zip", "latitude", "longitude", "Territory", "Secondary Territory",
                     "Distance to Marquee (mi)", "Closest Marquee"]
        drop_cols = [c for c in drop_cols if c in master_df.columns]
        master_cols = master_df.drop(columns=drop_cols).reset_index(drop=True)
        master_subset = master_cols.iloc[primary["_idx"].values].reset_index(drop=True)
        meta = primary[["Territory", "Secondary Territory", "Distance to Marquee (mi)", "Closest Marquee"]].reset_index(drop=True)
        result = pd.concat([meta, master_subset], axis=1)
        result = result.sort_values(["Territory", "Distance to Marquee (mi)"]).reset_index(drop=True)

    summary_rows = []
    for terr in territories:
        within = int((territory_dist[terr] <= max_distance_mi).sum())
        if not result.empty:
            terr_rows = result[result["Territory"] == terr]
            count = len(terr_rows)
            max_d = terr_rows["Distance to Marquee (mi)"].max() if count else None
            n_opco = int(terr_rows.get("OpCo?", pd.Series([], dtype=bool)).sum())
            n_active = int(terr_rows.get("Active <14d?", pd.Series([], dtype=bool)).sum())
            n_owner = int(terr_rows.get("Has Active Owner?", pd.Series([], dtype=bool)).sum())
        else:
            count = 0
            max_d = None
            n_opco = n_active = n_owner = 0
        summary_rows.append({
            "Territory": terr,
            "Count": count,
            "Target": target,
            "Shortfall?": "YES" if count < target else "",
            "ICPs in Range": within,
            "Max Distance (mi)": round(max_d, 1) if max_d is not None else None,
            "# OpCos": n_opco,
            f"# Active <{DEFAULT_ACTIVITY_DAYS}d": n_active,
            "# With Active Owner": n_owner,
        })
    summary_df = pd.DataFrame(summary_rows)
    return result, summary_df


def build_comparison_summary(results_by_cap: dict, target: int, primary_cap: int,
                             activity_days: int) -> pd.DataFrame:
    """Build a side-by-side cap-comparison summary.

    Columns: Territory, Target, Count @<cap>mi for each cap, Max Distance @<cap>mi for each cap,
             then ICPs in Range / # OpCos / # Active / # With Active Owner using the primary cap.
    """
    caps = sorted(results_by_cap.keys())
    primary_summary = results_by_cap[primary_cap][1]
    territories = primary_summary["Territory"].tolist()
    active_col = f"# Active <{activity_days}d"

    rows = []
    for terr in territories:
        row = {"Territory": terr, "Target": target}
        for cap in caps:
            sr = results_by_cap[cap][1]
            sub = sr[sr["Territory"] == terr].iloc[0]
            row[f"Count @{cap}mi"] = int(sub["Count"])
        for cap in caps:
            sr = results_by_cap[cap][1]
            sub = sr[sr["Territory"] == terr].iloc[0]
            row[f"Max Distance @{cap}mi"] = sub["Max Distance (mi)"]
        pr = primary_summary[primary_summary["Territory"] == terr].iloc[0]
        row[f"ICPs in Range @{primary_cap}mi"] = int(pr["ICPs in Range"])
        row["# OpCos"] = int(pr["# OpCos"])
        row[active_col] = int(pr[active_col])
        row["# With Active Owner"] = int(pr["# With Active Owner"])
        rows.append(row)
    return pd.DataFrame(rows)


def build_key_findings(comparison_summary: pd.DataFrame, target: int, primary_cap: int) -> list[str]:
    """Generate dynamic, data-driven findings text for the Summary tab.

    Stays factual — pulls numbers directly from the comparison rather than hardcoding.
    Findings only appear if they apply (no empty bullets).
    """
    caps = sorted(int(c.replace("Count @", "").replace("mi", ""))
                  for c in comparison_summary.columns if c.startswith("Count @"))
    primary_count_col = f"Count @{primary_cap}mi"
    largest_cap = max(caps)
    largest_count_col = f"Count @{largest_cap}mi"
    n_total = len(comparison_summary)

    findings: list[str] = []

    # Headline coverage at primary cap
    n_hit = int((comparison_summary[primary_count_col] >= target).sum())
    findings.append(
        f"{n_hit} of {n_total} territories hit {target} accounts at the @{primary_cap}mi (recommended) cap; "
        f"{n_total - n_hit} fall short."
    )

    # Persistent shortfalls — geography-bound (don't hit even at largest cap)
    persistent = comparison_summary[comparison_summary[largest_count_col] < target]
    if len(persistent):
        parts = ", ".join(
            f"{r['Territory']} maxes at {int(r[largest_count_col])}"
            for _, r in persistent.iterrows()
        )
        findings.append(
            f"Persistent shortfalls — geography prevents reaching {target} even at the @{largest_cap}mi cap: {parts}. "
            f"Consider pairing these pods with a neighbor for SDR motion."
        )

    # Stretched territories — hit target only above primary cap
    stretched = []
    for _, row in comparison_summary.iterrows():
        if row[primary_count_col] >= target or row[largest_count_col] < target:
            continue
        for cap in caps:
            if cap <= primary_cap:
                continue
            if row[f"Count @{cap}mi"] >= target:
                md = row[f"Max Distance @{cap}mi"]
                stretched.append(f"{row['Territory']} (needs @{cap}mi, max {md:.0f}mi)")
                break
    if stretched:
        findings.append(
            f"Hit {target} only above primary cap — pod becomes regional, not metro: {'; '.join(stretched)}."
        )

    # Density leaders — territories with deepest well at primary cap (info, not concern)
    range_col = f"ICPs in Range @{primary_cap}mi"
    if range_col in comparison_summary.columns:
        top3 = comparison_summary.nlargest(3, range_col)
        parts = ", ".join(f"{r['Territory']} ({int(r[range_col])})" for _, r in top3.iterrows())
        findings.append(
            f"Deepest wells at @{primary_cap}mi (most ICPs in range — room to scale beyond {target}): {parts}."
        )

    # Activity / ownership signal
    if "# With Active Owner" in comparison_summary.columns:
        most_owned = comparison_summary.nlargest(1, "# With Active Owner").iloc[0]
        if int(most_owned["# With Active Owner"]) >= 25:
            findings.append(
                f"Highest already-owned territory: {most_owned['Territory']} has {int(most_owned['# With Active Owner'])} "
                f"of {int(most_owned[primary_count_col])} accounts already assigned to a real rep — heavier coordination needed before SDR launch."
            )

    return findings


def build_definitions_df(activity_days: int, primary_cap: int) -> pd.DataFrame:
    """Two-column reference of every Summary column and what it means."""
    rows = [
        ("Territory", "Geographic pod / market name."),
        ("Target", "Target accounts per pod (default 100)."),
        ("Count @<cap>mi", "Number of classifier-confirmed ICP Fits assigned to this territory at the given max-distance cap. Each cap is a 'what if' — pick the cap that best balances pod size vs geographic stretch. Backfill is enabled, so each Count includes both the natural top-100 (after overlap resolution to the closest territory) and any backfilled accounts within the cap. NEVER additive across columns."),
        ("Max Distance @<cap>mi", "Under each cap, the distance from the territory's nearest marquee to the farthest account in the pod. Tells you how stretched the pod becomes at that cap. Higher = more outbound driving / less geographic coherence."),
        (f"ICPs in Range @{primary_cap}mi", f"Total classifier-confirmed ICP Fits within {primary_cap}mi of any marquee in this territory. NOT additive to Count — many of these were assigned to a closer territory. Read as: 'how deep is the well in this geography.'"),
        ("# OpCos", f"Of the @{primary_cap}mi pod: how many assigned accounts have a Parent Account ID set in SFDC (i.e., subsidiary of a larger parent). Surfaced for awareness — NOT removed from Count."),
        (f"# Active <{activity_days}d", f"Of the @{primary_cap}mi pod: how many have a Sales Last Activity Date within the last {activity_days} days. NOT removed from Count — surfaced so reps know which accounts are warm."),
        ("# With Active Owner", f"Of the @{primary_cap}mi pod: how many have an Account Owner that is NOT the SFDC 'Marketing User' placeholder (i.e., already assigned to a real rep). NOT removed from Count — surfaced so reps know which accounts are already worked."),
        ("Secondary Territory (Assigned Accounts sheet)", "Other territories that also reach this account within the primary cap. The account is assigned to its closest territory; secondaries are listed for awareness."),
        ("Assigned Accounts sheet", f"Per-account list of classifier-confirmed ICP Fits assigned to each territory at the primary cap (@{primary_cap}mi). This is the working list reps run. To regenerate at a different cap, re-run with --max-distance-miles <value>."),
    ]
    return pd.DataFrame(rows, columns=["Column", "Definition"])


def write_styled_output(output_path: Path, summary: pd.DataFrame, definitions: pd.DataFrame,
                        assigned_accounts: pd.DataFrame, territory_marquees: pd.DataFrame,
                        findings: list[str], target: int, primary_cap: int) -> None:
    """Write the four sheets and apply formatting.

    Summary layout (top-down):
      Row 1: 'KEY FINDINGS' title bar (dark blue, white text, merged across all columns)
      Row 2..K: one finding per row (light blue-gray fill, italic, merged across all columns)
      Row K+1: blank separator
      Row K+2: group banners ('Count by max-distance cap' / 'Max Distance by cap' / 'Subset of Count @<primary>mi')
      Row K+3: column headers
      Row K+4..N: data — Count cells where < Target get light-red fill + bold red text
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    GROUP_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    GROUP_FONT = Font(bold=True, italic=True, color="1F4E79", size=10)
    SHORTFALL_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    BOLD = Font(bold=True)
    BOLD_RED = Font(bold=True, color="9C0006")
    CENTER = Alignment(horizontal="center", vertical="center")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
    FINDING_FONT = Font(italic=True, color="1F4E79", size=11)

    with pd.ExcelWriter(output_path, engine="openpyxl") as w:
        # Findings block: 1 title row + N finding rows + 1 blank separator. Banner row sits at
        # findings_block_height + 1 (1-indexed), then header row follows, then data.
        n_findings = len(findings)
        findings_block_height = 1 + n_findings + 1  # title + findings + blank
        banner_row = findings_block_height + 1
        # pandas startrow is 0-indexed and points to the header row position
        summary_startrow = banner_row  # banner row above is at findings_block_height+1; header at +1; data at +2

        # We want: banner @ row banner_row, header @ row banner_row+1, data starting row banner_row+2.
        # pandas writes the column header at startrow+1 (1-indexed) when startrow is given.
        # So we set startrow=banner_row (0-indexed value = banner_row in 0-indexed = banner_row-1+1, hmm).
        # Cleaner: we manually wrote rows above; pass startrow=banner_row in 0-indexed
        # which puts pandas-written header at row banner_row+1 (1-indexed) and data at banner_row+2.
        summary.to_excel(w, sheet_name="Summary", index=False, startrow=banner_row)
        definitions.to_excel(w, sheet_name="Definitions", index=False)
        assigned_accounts.to_excel(w, sheet_name="Assigned Accounts", index=False)
        territory_marquees.to_excel(w, sheet_name="Territory Marquees", index=False)

        # ── Summary styling ──
        ws = w.sheets["Summary"]
        cols = list(summary.columns)
        n_cols = len(cols)
        n_rows = len(summary)

        # Findings title (row 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "KEY FINDINGS"
        title_cell.fill = HEADER_FILL
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER
        for c in range(1, n_cols + 1):
            ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.row_dimensions[1].height = 24

        # Findings rows (rows 2..1+n_findings)
        wrap_left = Alignment(wrap_text=True, horizontal="left", vertical="center", indent=1)
        for i, finding in enumerate(findings, start=1):
            r = 1 + i
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            cell = ws.cell(row=r, column=1)
            cell.value = "•  " + finding
            cell.fill = GROUP_FILL
            cell.font = FINDING_FONT
            cell.alignment = wrap_left
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = GROUP_FILL
            # row height roughly proportional to text length
            est_lines = max(1, len(finding) // 110 + 1)
            ws.row_dimensions[r].height = 18 * est_lines

        # Banner row (row banner_row in 1-indexed)
        count_col_idxs = [i for i, c in enumerate(cols) if c.startswith("Count @")]
        maxdist_col_idxs = [i for i, c in enumerate(cols) if c.startswith("Max Distance @")]
        flag_col_idxs = [i for i, c in enumerate(cols)
                         if c == "# OpCos" or c == "# With Active Owner" or c.startswith("# Active ")]

        def _banner(first_idx: int, last_idx: int, text: str) -> None:
            f, l = first_idx + 1, last_idx + 1
            ws.merge_cells(start_row=banner_row, start_column=f, end_row=banner_row, end_column=l)
            cell = ws.cell(row=banner_row, column=f)
            cell.value = text
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
            cell.alignment = CENTER
            for c in range(f, l + 1):
                ws.cell(row=banner_row, column=c).fill = GROUP_FILL

        if count_col_idxs:
            _banner(count_col_idxs[0], count_col_idxs[-1], "Count by max-distance cap")
        if maxdist_col_idxs:
            _banner(maxdist_col_idxs[0], maxdist_col_idxs[-1], "Max distance to farthest pod account")
        if flag_col_idxs:
            _banner(flag_col_idxs[0], flag_col_idxs[-1],
                    f"Subset of Count @{primary_cap}mi (not removed from Count)")

        # Header row (banner_row + 1)
        header_row = banner_row + 1
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        # Data rows
        first_data_row = header_row + 1
        for r in range(first_data_row, first_data_row + n_rows):
            for ci in count_col_idxs:
                cell = ws.cell(row=r, column=ci + 1)
                if isinstance(cell.value, (int, float)) and cell.value < target:
                    cell.fill = SHORTFALL_FILL
                    cell.font = BOLD_RED
                else:
                    cell.font = BOLD

        # Column widths
        for i, c in enumerate(cols, start=1):
            if c == "Territory":
                width_ = 24
            elif c == "Target":
                width_ = 9
            elif c.startswith("Count @"):
                width_ = 13
            elif c.startswith("Max Distance @"):
                width_ = 17
            elif c.startswith("ICPs in Range"):
                width_ = 22
            elif c == "# OpCos":
                width_ = 11
            elif c.startswith("# Active "):
                width_ = 16
            elif c == "# With Active Owner":
                width_ = 22
            else:
                width_ = 15
            ws.column_dimensions[get_column_letter(i)].width = width_

        # Freeze just below the header row
        ws.freeze_panes = ws.cell(row=first_data_row, column=1).coordinate

        # ── Definitions sheet: bold header, wider columns, wrap text ──
        ws2 = w.sheets["Definitions"]
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 110
        for c in range(1, 3):
            cell = ws2.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        wrap = Alignment(wrap_text=True, vertical="top")
        for r in range(2, len(definitions) + 2):
            ws2.cell(row=r, column=1).font = BOLD
            ws2.cell(row=r, column=2).alignment = wrap
        ws2.freeze_panes = "A2"

        # ── Assigned Accounts + Territory Marquees: just style the header row + freeze ──
        for sheet_name in ("Assigned Accounts", "Territory Marquees"):
            ws_other = w.sheets[sheet_name]
            ncol = ws_other.max_column
            for c in range(1, ncol + 1):
                cell = ws_other.cell(row=1, column=c)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER
            ws_other.freeze_panes = "A2"


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--marquees", default="inputs/marquees", help="Dir of per-territory marquee CSVs")
    p.add_argument("--master", required=True, help="Classified master xlsx path")
    p.add_argument("--enrich-master", default=None,
                   help="Optional fresh SFDC pull (CSV/xlsx) merged onto classified master by Account ID — adds Parent Account ID, Sales Last Activity Date, refreshes Account Owner")
    p.add_argument("--territory-mapping", default="inputs/territory_mapping.csv",
                   help="CSV mapping Account Name → Territory (overrides per-CSV territory and filename fallback)")
    p.add_argument("--output", default=None, help="Output xlsx (default outputs/assigned_<ts>.xlsx)")
    p.add_argument("--target", type=int, default=DEFAULT_TARGET)
    p.add_argument("--max-distance-miles", type=float, default=DEFAULT_PRIMARY_CAP_MI,
                   help=f"Primary cap (mi) used for Assigned Accounts and flag counts. Default {DEFAULT_PRIMARY_CAP_MI}. Summary shows comparison across {COMPARE_CAPS_MI} regardless.")
    p.add_argument("--activity-days", type=int, default=DEFAULT_ACTIVITY_DAYS,
                   help="Window (in days) for the recent-activity flag")
    p.add_argument("--include-all", action="store_true", help="Skip ICP Fit filter on master")
    args = p.parse_args()

    marquees_dir = Path(args.marquees)
    master_path = Path(args.master)
    mapping_path = Path(args.territory_mapping) if args.territory_mapping else None

    if not marquees_dir.is_dir():
        print(f"ERROR: marquees dir not found: {marquees_dir}", file=sys.stderr)
        return 2
    if not master_path.is_file():
        print(f"ERROR: master file not found: {master_path}", file=sys.stderr)
        return 2

    output_path = Path(args.output) if args.output else (
        Path("outputs") / f"assigned_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading marquees from {marquees_dir}...")
    marquees_df = load_marquees(marquees_dir, mapping_path)
    if marquees_df.empty:
        print("ERROR: no marquees loaded", file=sys.stderr)
        return 2
    print(f"  {len(marquees_df)} marquees across {marquees_df['Territory'].nunique()} territories")
    print(f"  Per-territory counts: " + ", ".join(
        f"{t}={n}" for t, n in marquees_df.groupby('Territory').size().items()
    ))

    print(f"Loading master from {master_path}...")
    master_df = pd.read_excel(master_path)
    print(f"  {len(master_df)} master rows")

    if args.enrich_master:
        enrich_path = Path(args.enrich_master)
        if not enrich_path.is_file():
            print(f"ERROR: enrich-master file not found: {enrich_path}", file=sys.stderr)
            return 2
        print(f"Merging enrichment fields from {enrich_path}...")
        before = len(master_df)
        master_df = merge_enrichment(master_df, enrich_path)
        added = [c for c in ("Parent Account ID", "Parent Account", "Sales Last Activity Date") if c in master_df.columns]
        print(f"  {len(master_df)} rows after merge (was {before}); added/refreshed: {', '.join(added) or 'none'}")

    print(f"Computing flags (activity window: {args.activity_days}d)...")
    master_df = compute_flags(master_df, activity_days=args.activity_days)

    if not args.include_all and "Updated Category" in master_df.columns:
        before = len(master_df)
        master_df = master_df[master_df["Updated Category"] == "ICP Fit"].reset_index(drop=True)
        print(f"  Filtered to ICP Fit only: {len(master_df)} (from {before})")

    primary_cap = int(args.max_distance_miles)
    compare_caps = sorted(set(COMPARE_CAPS_MI) | {primary_cap})

    print(f"Assigning (target={args.target}/territory) at caps {compare_caps}mi (primary={primary_cap}mi)...")
    results_by_cap = {}
    t0 = time.time()
    for cap in compare_caps:
        result, summary = assign(master_df, marquees_df, args.target, cap)
        results_by_cap[cap] = (result, summary)
        print(f"  cap={cap}mi: {len(result)} accounts assigned")
    print(f"  Total assignment time: {time.time() - t0:.1f}s")

    primary_result, _ = results_by_cap[primary_cap]
    comparison_summary = build_comparison_summary(results_by_cap, args.target, primary_cap, args.activity_days)
    findings = build_key_findings(comparison_summary, args.target, primary_cap)

    territory_marquees = (marquees_df[["Territory", "Marquee Account Name", "Marquee ZIP"]]
                          .sort_values(["Territory", "Marquee Account Name"])
                          .reset_index(drop=True))

    definitions = build_definitions_df(args.activity_days, primary_cap)

    write_styled_output(output_path, comparison_summary, definitions, primary_result,
                        territory_marquees, findings, args.target, primary_cap)

    print()
    print("=== Key Findings ===")
    for f in findings:
        print(f"  • {f}")
    print()
    print("=== Summary (cap comparison) ===")
    print(comparison_summary.to_string(index=False))
    print()
    print(f"Output: {output_path}")
    print(f"Assigned Accounts sheet uses primary cap = {primary_cap}mi")
    return 0


if __name__ == "__main__":
    sys.exit(main())
