"""Build a 6-sheet territory planning report from a classified xlsx.

Usage:
    python build_territory_report.py \\
        --input runs/all_territories_20260421_classified.xlsx \\
        --output runs/all_territories_20260421_territory_report.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Density thresholds (per MM Territory Planning Project plan: 250-350 total per AE across 2 geos → 125-175 per geo)
DENSITY_SUFFICIENT = 175
DENSITY_MARGINAL = 100

REV_LARGE = 50_000_000
REV_MID = 25_000_000

NAICS_ELECTRICAL = "238210"
NAICS_MECH_PLUMBING = "238220"

# Reviewer assignments for the 20-account proxy pass (per 2026-04-21 call).
# Jared gets Phoenix (AZ) + Atlanta (GA); Vincent keeps SA-Austin + takes DC-Baltimore;
# Jeremy = DFW + Houston; Sadie = Tampa-FL + Nashville-Charlotte; Jaime = NYC-NJ + Minneapolis-Louisville.
# Other TX is intentionally unassigned for now.
REVIEWER_ASSIGNMENTS: dict[str, str] = {
    "Phoenix": "Jared",
    "Atlanta": "Jared",
    "SA-Austin": "Vincent",
    "DC - Baltimore": "Vincent",
    "DFW": "Jeremy",
    "Houston": "Jeremy",
    "Tampa - FL": "Sadie",
    "Nashville - Charlotte": "Sadie",
    "NYC - NJ": "Jaime",
    "Minneapolis - Louisville": "Jaime",
    "Other TX": "(unassigned)",
}
PROXY_SAMPLE_SIZE = 20
PROXY_SAMPLE_SEED = 2026_04_21


# Colors
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def _norm_prior_category(v: object) -> str:
    """Normalize Prior Category values so 'ICP Fit (pattern-verified)' and '(Confirmed)' both map to 'ICP Fit'."""
    if pd.isna(v):
        return "(blank)"
    s = str(v).strip()
    if s.startswith("ICP Fit"):
        return "ICP Fit"
    return s


def _rev(v: object) -> int | None:
    if pd.isna(v) or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _density_verdict(icp_fit_count: int) -> tuple[str, PatternFill]:
    if icp_fit_count >= DENSITY_SUFFICIENT:
        return "Sufficient", GREEN
    if icp_fit_count >= DENSITY_MARGINAL:
        return "Marginal", YELLOW
    return "Combine or expand", RED


def build_summary_rows(df: pd.DataFrame) -> list[dict]:
    rows: list[dict] = []
    # Include "TOTAL" as a synthetic territory
    territories = list(df["Territory"].dropna().unique()) + ["TOTAL"]

    for terr in territories:
        sub = df if terr == "TOTAL" else df[df["Territory"] == terr]

        total = len(sub)
        script_cats = sub["Updated Category"].value_counts(dropna=False)
        prior_cats = sub["Prior Category"].map(_norm_prior_category).value_counts(dropna=False)

        script_fit = int(script_cats.get("ICP Fit", 0))
        script_review = int(script_cats.get("Needs Human Review", 0))
        script_unlikely = int(script_cats.get("Unlikely ICP", 0))
        prior_fit = int(prior_cats.get("ICP Fit", 0))
        delta = script_fit - prior_fit

        fit_rows = sub[sub["Updated Category"] == "ICP Fit"]
        large = sum(1 for v in fit_rows["Annual Revenue"] if (_rev(v) or 0) >= REV_LARGE)
        mid = sum(1 for v in fit_rows["Annual Revenue"] if REV_MID <= (_rev(v) or 0) < REV_LARGE)
        small = script_fit - large - mid

        naics_series = fit_rows["NAICS Code"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        elec = int((naics_series == NAICS_ELECTRICAL).sum())
        mech = int((naics_series == NAICS_MECH_PLUMBING).sum())
        other_naics = script_fit - elec - mech

        verdict, fill = _density_verdict(script_fit)

        rows.append({
            "Territory": terr,
            "Total Accounts": total,
            "Script ICP Fit": script_fit,
            "Script Needs Review": script_review,
            "Script Unlikely": script_unlikely,
            "Prior ICP Fit": prior_fit,
            "Δ Script vs Prior": delta,
            "Large ICP ($50M+)": large,
            "Mid ICP ($25-50M)": mid,
            "Small ICP (<$25M)": small,
            "Electrical NAICS (238210)": elec,
            "Mech/Plumbing NAICS (238220)": mech,
            "Other NAICS": other_naics,
            "Density Verdict": verdict,
            "_verdict_fill": fill,
            "_is_total": terr == "TOTAL",
        })
    return rows


def write_summary_sheet(wb: Workbook, summary_rows: list[dict]) -> None:
    ws = wb.create_sheet("Territory Planning Summary")

    # Omit metadata fields from the visible row
    visible_cols = [k for k in summary_rows[0].keys() if not k.startswith("_")]

    # Header
    for c, header in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for r_idx, row in enumerate(summary_rows, start=2):
        for c_idx, col in enumerate(visible_cols, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            if col == "Density Verdict":
                cell.fill = row["_verdict_fill"]
                cell.font = BOLD
            if row["_is_total"]:
                cell.font = BOLD
        # Trade-mix dominant bold: whichever of elec / mech / other is highest gets bold
        elec_val = row["Electrical NAICS (238210)"]
        mech_val = row["Mech/Plumbing NAICS (238220)"]
        other_val = row["Other NAICS"]
        m = max(elec_val, mech_val, other_val)
        if m > 0:
            if elec_val == m:
                ws.cell(row=r_idx, column=visible_cols.index("Electrical NAICS (238210)") + 1).font = BOLD
            if mech_val == m:
                ws.cell(row=r_idx, column=visible_cols.index("Mech/Plumbing NAICS (238220)") + 1).font = BOLD
            if other_val == m:
                ws.cell(row=r_idx, column=visible_cols.index("Other NAICS") + 1).font = BOLD

    # Column widths
    widths = {
        "Territory": 18, "Total Accounts": 14, "Script ICP Fit": 15,
        "Script Needs Review": 20, "Script Unlikely": 16, "Prior ICP Fit": 14,
        "Δ Script vs Prior": 18, "Large ICP ($50M+)": 18, "Mid ICP ($25-50M)": 18,
        "Small ICP (<$25M)": 18, "Electrical NAICS (238210)": 22,
        "Mech/Plumbing NAICS (238220)": 24, "Other NAICS": 14,
        "Density Verdict": 20,
    }
    for c_idx, col in enumerate(visible_cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 14)
    ws.row_dimensions[1].height = 44
    ws.freeze_panes = "B2"


def write_changes_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Changes")

    df = df.copy()
    df["Prior (normalized)"] = df["Prior Category"].map(_norm_prior_category)
    changes = df[df["Updated Category"] != df["Prior (normalized)"]].copy()

    cols = ["Territory", "Account Name", "Website",
            "Prior Category", "Updated Category", "Website Evidence",
            "Prior Evidence", "Annual Revenue", "NAICS Code",
            "Account Owner", "Salesforce URL", "Account ID (18 Char)"]
    changes = changes[[c for c in cols if c in changes.columns]].rename(columns={
        "Updated Category": "Script Category",
        "Website Evidence": "Script Evidence",
        "NAICS Code": "NAICS",
    })
    changes = changes.sort_values(by=["Territory", "Script Category", "Account Name"]).reset_index(drop=True)

    # Write header
    for c_idx, col in enumerate(changes.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(changes.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))

    widths = {"Territory": 14, "Account Name": 32, "Website": 30,
              "Prior Category": 24, "Script Category": 20, "Script Evidence": 60,
              "Prior Evidence": 60, "Annual Revenue": 14, "NAICS": 10,
              "Account Owner": 18, "Salesforce URL": 30, "Account ID (18 Char)": 22}
    for c_idx, col in enumerate(changes.columns, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 18)
    ws.freeze_panes = "A2"


def write_target_accounts_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Target Accounts (ICP Fit)")

    fits = df[df["Updated Category"] == "ICP Fit"].copy()
    fits["Annual Revenue (num)"] = fits["Annual Revenue"].map(_rev)
    fits["Owner Gap"] = fits["Account Owner"].fillna("").astype(str).str.strip().isin(["", "Marketing User"])
    fits = fits.sort_values(
        by=["Territory", "Annual Revenue (num)"],
        ascending=[True, False],
        na_position="last",
    ).reset_index(drop=True)

    cols_map = [
        ("Territory", "Territory"),
        ("Account Name", "Account Name"),
        ("Annual Revenue", "Annual Revenue"),
        ("Website", "Website"),
        ("Website Evidence", "Script Evidence"),
        ("NAICS Code", "NAICS"),
        ("State", "State"),
        ("ZIP", "ZIP"),
        ("Account Type", "Account Type"),
        ("Account Owner", "Account Owner"),
        ("SDR Owner", "SDR Owner"),
        ("Owner Gap", "Owner Gap"),
        ("Salesforce URL", "Salesforce URL"),
        ("Account ID (18 Char)", "Account ID"),
    ]
    existing = [(src, disp) for src, disp in cols_map if src in fits.columns]

    for c_idx, (_src, disp) in enumerate(existing, start=1):
        cell = ws.cell(row=1, column=c_idx, value=disp)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(fits.iterrows(), start=2):
        for c_idx, (src, _) in enumerate(existing, start=1):
            val = row[src]
            cell = ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))
            if src == "Owner Gap" and bool(val):
                cell.fill = YELLOW

    widths = {"Territory": 14, "Account Name": 32, "Annual Revenue": 14,
              "Website": 30, "Script Evidence": 60, "NAICS": 10, "State": 8,
              "ZIP": 10, "Account Type": 14, "Account Owner": 18, "SDR Owner": 18,
              "Owner Gap": 10, "Salesforce URL": 30, "Account ID": 22}
    for c_idx, (_, disp) in enumerate(existing, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(disp, 18)
    ws.freeze_panes = "A2"


def write_all_accounts_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("All Accounts")
    ordered_cols = [
        "Territory", "Account Name", "Website",
        "Prior Category", "Updated Category", "Prior Evidence", "Website Evidence",
        "Annual Revenue", "NAICS Code", "Account Type", "State", "ZIP", "Address",
        "Account Owner", "SDR Owner",
        "Verification Method", "Scrape Timestamp",
        "Salesforce URL", "Account ID (18 Char)",
    ]
    cols = [c for c in ordered_cols if c in df.columns]
    display = df[cols].rename(columns={
        "Updated Category": "Script Category",
        "Website Evidence": "Script Evidence",
    })

    for c_idx, col in enumerate(display.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(display.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))

    ws.freeze_panes = "A2"


def write_proxy_samples_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sample PROXY_SAMPLE_SIZE rows per territory from the Needs Human Review bucket.
    Used by reviewers (per REVIEWER_ASSIGNMENTS) to establish a proxy percentage —
    what fraction of Needs Review accounts are actually ICP Fit on manual inspection.
    """
    ws = wb.create_sheet("Proxy Samples")

    needs_review = df[df["Updated Category"] == "Needs Human Review"].copy()
    samples: list[pd.DataFrame] = []
    for terr in df["Territory"].dropna().unique():
        sub = needs_review[needs_review["Territory"] == terr]
        if len(sub) == 0:
            continue
        n = min(PROXY_SAMPLE_SIZE, len(sub))
        samples.append(sub.sample(n=n, random_state=PROXY_SAMPLE_SEED))

    if not samples:
        ws.cell(row=1, column=1, value="No Needs Human Review rows to sample.")
        return

    combined = pd.concat(samples, ignore_index=True)
    combined["Reviewer"] = combined["Territory"].map(REVIEWER_ASSIGNMENTS).fillna("(unassigned)")
    combined["Proxy Verdict"] = ""  # Blank column for reviewer to fill in
    combined["Reviewer Notes"] = ""

    combined = combined.sort_values(by=["Reviewer", "Territory", "Account Name"]).reset_index(drop=True)

    display_cols = [
        ("Reviewer", "Reviewer"),
        ("Territory", "Territory"),
        ("Account Name", "Account Name"),
        ("Website", "Website"),
        ("Website Evidence", "Script Evidence (why flagged)"),
        ("Proxy Verdict", "Proxy Verdict (ICP Fit / Unlikely ICP / Still Unclear)"),
        ("Reviewer Notes", "Reviewer Notes"),
        ("Annual Revenue", "Annual Revenue"),
        ("NAICS Code", "NAICS"),
        ("Account Type", "Account Type"),
        ("State", "State"),
        ("ZIP", "ZIP"),
        ("Account Owner", "Account Owner"),
        ("Salesforce URL", "Salesforce URL"),
        ("Account ID (18 Char)", "Account ID"),
    ]
    existing = [(src, disp) for src, disp in display_cols if src in combined.columns or src in ("Proxy Verdict", "Reviewer", "Reviewer Notes")]

    for c_idx, (_src, disp) in enumerate(existing, start=1):
        cell = ws.cell(row=1, column=c_idx, value=disp)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(combined.iterrows(), start=2):
        for c_idx, (src, _) in enumerate(existing, start=1):
            val = row[src] if src in row.index else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))
            if src == "Reviewer" and val == "(unassigned)":
                cell.fill = YELLOW

    widths = {"Reviewer": 12, "Territory": 18, "Account Name": 32,
              "Website": 30, "Script Evidence (why flagged)": 60,
              "Proxy Verdict (ICP Fit / Unlikely ICP / Still Unclear)": 45,
              "Reviewer Notes": 30, "Annual Revenue": 14, "NAICS": 10,
              "Account Type": 14, "State": 8, "ZIP": 10, "Account Owner": 18,
              "Salesforce URL": 30, "Account ID": 22}
    for c_idx, (_, disp) in enumerate(existing, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(disp, 18)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 44


def write_methodology_sheet(wb: Workbook, df: pd.DataFrame, meta: dict) -> None:
    ws = wb.create_sheet("Methodology")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100

    lines = [
        ("Kojo ICP Classifier — Phase A", ""),
        ("", ""),
        ("Run timestamp", meta.get("run_timestamp", "")),
        ("Classifier script path", "pipeline-pulse-app/scripts/icp-classifier/classify.py"),
        ("Classifier model", meta.get("model", "claude-haiku-4-5-20251001")),
        ("Max workers (concurrency)", str(meta.get("max_workers", 1))),
        ("Per-domain scrape delay (s)", str(meta.get("per_domain_delay", 1.5))),
        ("Scraped text max chars", str(meta.get("max_text_chars", 5000))),
        ("Source xlsx", meta.get("source_xlsx", "")),
        ("Territories in this run", meta.get("territories", "")),
        ("Total rows classified", str(len(df))),
        ("", ""),
        ("Columns added by this tool",
            "Script Category (ICP Fit / Needs Human Review / Unlikely ICP), Script Evidence (one-line rationale), "
            "Verification Method (Website scrape / Website inaccessible / Classifier parse error), Scrape Timestamp (UTC ISO 8601)"),
        ("Prior Category preserved",
            "Yes — Prior Category and Prior Evidence columns retain the original xlsx values (NAICS pattern-matched "
            "OR Jaime Claude Desktop verification). Nothing in the source xlsx is overwritten."),
        ("Density thresholds used",
            f"Sufficient ≥{DENSITY_SUFFICIENT} ICP Fit per territory; Marginal {DENSITY_MARGINAL}–{DENSITY_SUFFICIENT-1}; "
            f"<{DENSITY_MARGINAL} Combine or expand. Derived from MM Territory Planning Project "
            "(target 250–350 accounts per AE across 2 geos)."),
        ("Revenue tiers",
            f"Large: ≥${REV_LARGE:,} | Mid: ${REV_MID:,}–${REV_LARGE:,} | Small: <${REV_MID:,}"),
        ("Trade mix NAICS buckets",
            f"Electrical: {NAICS_ELECTRICAL} | Mech/Plumbing: {NAICS_MECH_PLUMBING} | Other: everything else "
            "(includes 238110 plumbing/heating/AC, 238290 other building equipment)"),
        ("", ""),
        ("Related docs",
            "Design spec: docs/superpowers/specs/2026-04-21-icp-classifier-design.md | "
            "Implementation plan: docs/superpowers/plans/2026-04-21-icp-classifier.md | "
            "Notion project page: https://www.notion.so/349ae6b4a3b981589274e765924c0e05"),
        ("Validation note (2026-04-21)",
            "86-row Phoenix sample showed 94% agreement with Jaime Stillwell's manual Claude Desktop verification (16/17 rows). "
            "The 1 disagreement was DP Air Corp — Jaime 'Maybe?', script 'ICP Fit' with strong data-center evidence — arguably "
            "the script was more confident than the human, not wrong."),
    ]
    for r_idx, (label, value) in enumerate(lines, start=1):
        a = ws.cell(row=r_idx, column=1, value=label)
        b = ws.cell(row=r_idx, column=2, value=value)
        a.font = BOLD
        b.alignment = Alignment(wrap_text=True, vertical="top")
        if label.startswith("Kojo ICP"):
            a.font = Font(bold=True, size=14)


def build_report(input_path: Path, output_path: Path, meta: dict) -> None:
    df = pd.read_excel(input_path, dtype=object)

    wb = Workbook()
    # Remove the default 'Sheet'
    default = wb.active
    wb.remove(default)

    summary_rows = build_summary_rows(df)
    write_summary_sheet(wb, summary_rows)
    write_changes_sheet(wb, df)
    write_target_accounts_sheet(wb, df)
    write_proxy_samples_sheet(wb, df)
    write_all_accounts_sheet(wb, df)
    write_methodology_sheet(wb, df, meta)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Report written: {output_path}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Path to classified xlsx (with Prior + Updated columns)")
    p.add_argument("--output", required=True, help="Path to the territory report xlsx")
    p.add_argument("--source-xlsx", default="", help="Path to the original SFDC xlsx (for methodology metadata)")
    p.add_argument("--territories", default="", help="Comma-separated territory names (for methodology metadata)")
    args = p.parse_args()

    meta = {
        "run_timestamp": datetime.now().isoformat(timespec="seconds"),
        "model": "claude-haiku-4-5-20251001",
        "max_workers": 1,
        "per_domain_delay": 1.5,
        "max_text_chars": 5000,
        "source_xlsx": args.source_xlsx,
        "territories": args.territories,
    }
    build_report(Path(args.input), Path(args.output), meta)


if __name__ == "__main__":
    main()
